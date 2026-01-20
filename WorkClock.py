"""
WorkClock is a script to manage clocking in and out on for DESY Excel timesheet.
It assumes your file is named "/Zeiterfassungstabelle 2025 Doktorand_innen.xlsx" and that it is
in your home directory "~/"

It uses openpyxl to manage the Excel sheet 
(pip install openpyxl (my current version is 3.1.2) )

Arguments:
 action:
    - in    : to clock in
    - out   : to clock out
    - show  : to inspect the date's clock ins and outs
    - set   : to manually set a clock in and out
    - clear : to remove the date's clock ins and outs
    - fake  : to fill the sheet with clock ins and outs

 --shift: An optional integer to adjust the clock-in or clock-out time in minutes to the current time.

 --date: Specify the date, if unspecified takes today's date [format: DD/MM/YYYY]

 If action is "set":
     --cin:  specify the clock in time [format: HH:mm]

     --cout: specify the clock out time [format: HH:mm]

 If action is "fake":
     --startdate: specify the start date [format: DD/MM/YYYY] 
     
     --inmu: Mean time of morning's clock in (Default: 9:00)

     --instd: Std of morning's clock in 

     --lunchstartmu: Mean time of lunch start (Default: 12:30)

     --lunchstartstd: Std of lunch start 
     
     --lunchdurationmu: Mean time of lunch duration in minutes (Default: 30)

     --lunchduratiostd: Std of lunch duration

     --workdurationmu: Mean time of work duration [format: HH:mm] (Default : 8:00)

     --workdurationstd: Std of work duration

Examples:
 ❯ python WorkClock.py in          (to clock-in at the current time)
 ❯ python WorkClock.py out -s 15   (to clock-out 15 minutes after the current time)
 ❯ python WorkClock.py show        (show today's clock-in and clock-out times)
 ❯ python WorkClock.py clear       (clear today's clock-ins and outs)
 ❯ python WorkClock.py set --cin 11:00 --cout 12:30 (Clock in in today's date at 11:00 and clock out at 12:30)

For any command you can change the date with the --date flag (format DD/MM/YYYY), if not provided, it takes
today's date

Examples: 
 ❯ python3 WorkClock.py in --date 18/05/2025
     Clocked in at Sun, 18/05/2025 : 18:49
 ❯ python3 WorkClock.py clear --date 18/05/2025
     Day Sun, 18/05/2025 cleared
"""

import argparse
import os
from datetime import datetime, timedelta

import openpyxl as xl


def read_cell(cell):
    value = ':'.join(str(cell.value).split(":")[:2])
    return datetime.strptime(value, "%H:%M")

def clock_in(sheet, row, time_shift=0):
    """
    Record a clock-in time on the specified row.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
        time_shift (int, optional): Minutes to adjust the clock-in time. Defaults to 0.
    """
    for start_col, end_col in zip(clock_in_cols, clock_out_cols):
        start_cell = sheet.cell(row=row, column=start_col)
        end_cell = sheet.cell(row=row, column=end_col)

        date_cell = sheet.cell(row=row, column=4)
        cell_date = parse_date_cell(date_cell)

        if not start_cell.value and not end_cell.value:
            if start_col != clock_in_cols[0]:
                prev_end_cell = sheet.cell(row=row, column=clock_out_cols[clock_in_cols.index(start_col) - 1])
                if not validate_time_order(prev_end_cell, time_shift):
                    print(f"Error: Clock-in time must be after previous clock-out on {cell_date}")
                    return
            start_cell.value = get_adjusted_time(time_shift).time()
            start_cell.number_format = "HH:mm"
            start_cell_str = start_cell.value.strftime('%H:%M')
            print(f"Clocked in at {cell_date} : {start_cell_str}")
            return
        elif start_cell.value and not end_cell.value:
            print(f"Error: Clock out first on {cell_date}")
            return

    print(f"Error: No available clock-in column for {cell_date}")

def clock_out(sheet, row, time_shift=0):
    """
    Record a clock-out time on the specified row.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
        time_shift (int, optional): Minutes to adjust the clock-out time. Defaults to 0.
    """
    for start_col, end_col in zip(clock_in_cols, clock_out_cols):
        start_cell = sheet.cell(row=row, column=start_col)
        end_cell = sheet.cell(row=row, column=end_col)

        date_cell = sheet.cell(row=row, column=4)
        cell_date = parse_date_cell(date_cell)

        if start_cell.value and not end_cell.value:
            if not validate_time_order(start_cell, time_shift):
                print(f"Error: Clock-out time must be after clock-in on {cell_date}")
                return
            end_cell.value = get_adjusted_time(time_shift).time()
            end_cell.number_format = "HH:mm"
            end_cell_str = end_cell.value.strftime('%H:%M')
            print(f"Clocked out at {cell_date} : {end_cell_str}")
            return
        elif not start_cell.value:
            print(f"Error: Clock in first on {cell_date}")
            return

    print(f"Error: No available clock-out column for {cell_date}")

def calculate_total_time(sheet, row):
    """
    Calculate total working time for the specified date row.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
    """
    total_time = timedelta()

    for start_col, end_col in zip(clock_in_cols, clock_out_cols):
        start_cell = sheet.cell(row=row, column=start_col)
        end_cell = sheet.cell(row=row, column=end_col)

        if start_cell.value and end_cell.value:
            start_time = read_cell(start_cell)
            end_time = read_cell(end_cell)
            total_time += end_time - start_time

    return total_time

def inspect_row(sheet, row):
    """
    Inspect and print clock-in and out times for a specified date row.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
    """
    date_cell = sheet.cell(row=row, column=4)
    cell_date = parse_date_cell(date_cell)
    print(f"{cell_date}:")

    final_clockout = True
    last_clockin = None
    for index, (start_col, end_col) in enumerate(zip(clock_in_cols, clock_out_cols)):
        start_cell = sheet.cell(row=row, column=start_col)
        end_cell = sheet.cell(row=row, column=end_col)

        if start_cell.value:
            last_clockin = start_cell.value
            start_time = start_cell.value
            start_time_str = start_time.strftime('%H:%M')
            if end_cell.value:
                end_time = end_cell.value
                end_time_str = end_time.strftime('%H:%M')
            else:
                end_time_str = "???"
                final_clockout = False

            print(f"{index + 1}: {start_time_str} -> {end_time_str}")
            
    total_time = calculate_total_time(sheet, row)

    if not final_clockout:
        current_time = datetime.now().time()
        delta_time   = datetime.combine(date.today(), current_time) - datetime.combine(date.today(), last_clockin)
        if delta_time.total_seconds() > 0:
            total_time  += delta_time

    # Get total seconds
    total_seconds = int(total_time.total_seconds())

    # Compute hours and minutes
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    
    total_time = f"{hours:02} hours {minutes:02} minutes"

    print(f"Total time: {total_time}")

def parse_date_cell(cell):
    if isinstance(cell.value, datetime):
        return cell.value.strftime("%a, %d/%m/%Y")
    else:
        try:
            return datetime.strptime(cell.value, "%a, %d/%m/%Y").strftime("%a, %d/%m/%Y")
        except Exception:
            return "Unknown Date"

def validate_time_order(cell, time_shift):
    prev_time = read_cell(cell)
    return get_adjusted_time(time_shift) > prev_time

def get_adjusted_time(shift):
    return datetime.strptime(datetime.now().strftime("%H:%M"), "%H:%M") + timedelta(minutes=shift)

def clear_row(sheet, row):
    """
    Clear all clock-in and clock-out times for the specified row.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
    """
    for start_col, end_col in zip(clock_in_cols, clock_out_cols):
        sheet.cell(row=row, column=start_col).value = None
        sheet.cell(row=row, column=end_col).value = None
        
    date_cell = sheet.cell(row=row, column=4)
    cell_date = parse_date_cell(date_cell)
    print(f"Day {cell_date} cleared")

def set_time(sheet, row, start_time_str, end_time_str):
    """
    Manually set clock-in and clock-out times by calculating the time shift.

    Args:
        sheet (Worksheet): The Excel sheet.
        row (int): The row number.
        start_time_str (str): Clock-in time in "HH:mm" format.
        end_time_str (str): Clock-out time in "HH:mm" format.
    """
    current_time = datetime.now()

    start_time = datetime.strptime(start_time_str, "%H:%M").replace(year=current_time.year, month=current_time.month, day=current_time.day)
    start_shift = int((start_time - current_time).total_seconds() // 60) + 1
    clock_in(sheet, row, time_shift=start_shift)

    if end_time_str:
        end_time = datetime.strptime(end_time_str, "%H:%M").replace(year=current_time.year, month=current_time.month, day=current_time.day)
        end_shift = int((end_time - current_time).total_seconds() // 60) + 1
        clock_out(sheet, row, time_shift=end_shift)
        
def find_date(sheet, _date):
    for row in sheet.iter_rows(min_row=1, max_col=4, min_col=4):
        cell = row[0]
        try:
            cell_date = cell.value.date() if isinstance(cell.value, datetime) else datetime.strptime(cell.value, "%a, %d/%m/%Y").date()
        except Exception:
            continue

        if cell_date == _date:
            return cell_date
    else:
        print("Today's date not found.")
        exit(1)


def fake(sheet, start_date, end_date, in_mu, in_std, break_mu, break_std, worklen_mu, worklen_std):
    date_row_start = find_date(sheet, start_date)
    date_row_end   = find_date(sheet, end_date)

    def is_weekend(_date):
        """Return True if it is a weekend"""
        return _date.weekday() >= 5

    def is_date_clockable(_date):
        return not is_weekend(_date) and (not sheet.cell(row=date_row_end.row, column=5).value or not sheet.cell(row=date_row_end.row, column=5).value.isalpha())

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Manage time clocking on an Excel timesheet.")
    parser.add_argument("action", choices=["in", "out", "show", "set", "clear"], help="Choose to clock in, out, or inspect a row")
    parser.add_argument("--shift", type=int, default=0, help="Time shift in minutes (positive or negative)")
    parser.add_argument("--date", type=str, help="Date in format 'DD/MM/YYYY'")
    parser.add_argument("--cin",  type=str, help="Clock-in in format 'HH:mm'")
    parser.add_argument("--cout", type=str, help="Clock-out in format 'HH:mm'", default='')

    args = parser.parse_args()

    date = datetime.today().date()
    if args.date:
        try:
            date = datetime.strptime(args.date, "%d/%m/%Y").date()
        except ValueError:
            print("Invalid date format. Use: 'DD/MM/YYYY'")
            exit(1)

    clock_in_cols = [7, 9, 11, 13]
    clock_out_cols = [8, 10, 12, 14]
    date_start_row = 133
    date_col = 4
    additional_information_col = 6

    filename = os.path.expanduser("~/Zeiterfassungstabelle 2025 Doktorand_innen.xlsx")
    workbook = xl.load_workbook(filename)
    sheet = workbook['Time Recording']
    
    date_row = find_date(sheet, date)

    if args.action == "in":
        clock_in(sheet, date_row, time_shift=args.shift)
    elif args.action == "out":
        clock_out(sheet, date_row, time_shift=args.shift)
    elif args.action == "show":
        inspect_row(sheet, date_row)
    elif args.action == "set":
        set_time(sheet, date_row, args.cin, args.cout)
    elif args.action == "clear":
        clear_row(sheet, date_row)

    workbook.save(filename)


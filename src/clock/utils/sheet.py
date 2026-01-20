import warnings
from datetime import date, datetime, timedelta

import openpyxl as xl
from tabulate import tabulate

from ..utils.print import error, info, warn
from . import cell
from .file import EXCEL_FILE
from .time import format_timedelta

warnings.filterwarnings("ignore", message="DrawingML support is incomplete")


class Sheet:
    def __init__(self):
        self.SHEET_NAME = 'Time Recording'
        self.DATE_COLUMN = 4
        self.COMMENT_COLUMN = 6
        self.HEADER_ROW = 5
        self.WORK_COLUMNS = [
            {"start" : 7, "end" : 8},
            {"start" : 9, "end" : 10},
            {"start" : 11, "end" : 12},
            {"start" : 13, "end" : 14},
        ]

        self.workbook = xl.load_workbook(EXCEL_FILE)
        self.sheet = self.workbook[self.SHEET_NAME]

        self.DATE_ROW_FIRST = self.get_DATE_ROW_FIRST()
        
    def get_DATE_ROW_FIRST(self, debug: bool = False) -> int:
        row = 1

        while row <= self.sheet.max_row:
            c = self.sheet.cell(row=row, column=self.DATE_COLUMN)

            if cell.is_date(c):
                break

            if debug:
                info(f"Skipping row {row} | {c.value}")

            row += 1

        if debug:
            info(f"Date row found at {row} | {c.value}")

        return row
    
    def get_work_hours(self, row, verbose = False):
        total_time = timedelta(0)
        work_list = []
        empty_block_encountered = False

        date_target = self.sheet.cell(row=row, column=self.DATE_COLUMN).value
        for i, work_cols in enumerate(self.WORK_COLUMNS):
            work_start = self.sheet.cell(row=row, column=work_cols["start"]).value
            work_end = self.sheet.cell(row=row, column=work_cols["end"]).value

            # Normalize None and datetime
            # If values are strings or numbers, you might need parsing here
            # Assuming they are datetime.time or datetime.datetime or None
            
            if work_start is None and work_end is None:
                # Mark we have an empty pair
                empty_block_encountered = True
                continue

            # If empty_block_encountered before, but now we found filled pair → error (gap)
            if empty_block_encountered:
                if verbose:
                    error(f"({date_target.strftime('%d/%m/%Y')}) Gap detected: work period {i+1} has entries after empty pair")
                continue

            # Only end filled → error
            if work_start is None and work_end is not None:
                if verbose:
                    error(f"({date_target.strftime('%d/%m/%Y')}) Work start must be filled if work end is filled (pair {i+1})")
                continue

            # If only start filled (no end), consider incomplete — could raise or ignore
            if work_start is not None and work_end is None:
                if verbose:
                    warn(f"({date_target.strftime('%d/%m/%Y')}) Work end missing for work start (pair {i+1})")
                work_list.append({"start" : work_start, "end" : None})
                continue

            if datetime.combine(date.today(), work_end) < datetime.combine(date.today(), work_start):
                if verbose:
                    error(f"({date_target.strftime('%d/%m/%Y')}) Work end time is before start time (pair {i+1})")
                continue

            total_time += (datetime.combine(date.today(), work_end) - datetime.combine(date.today(), work_start))
            work_list.append({"start" : work_start, "end" : work_end})

        return total_time, work_list

    def date_to_row(self, target_date: date) -> int:
        row = self.DATE_ROW_FIRST

        while row <= self.sheet.max_row:
            c = self.sheet.cell(row=row, column=self.DATE_COLUMN)

            if isinstance(c.value, datetime):
                cell_date = c.value.date()
            else:
                cell_date = c.value

            if cell_date == target_date:
                return row

            row += 1

        error(f"Date {target_date} not found in column {self.DATE_COLUMN}")
        return 0
    
    def print_row(self, row: int):
        headers = ["Date", "Comments"]

        date_str  = self.sheet.cell(row=row, column=self.DATE_COLUMN).value.strftime("%d/%m/%Y")  # 15/01/2026
        comment_str = self.sheet.cell(row=row, column=self.COMMENT_COLUMN).value
        
        row_content = [[date_str, comment_str]]    
        
        # add work hours
        total_time, work_list = self.get_work_hours(row, verbose=True)
        if work_list:
            headers += ["Total time"]
            row_content[0] += [format_timedelta(total_time)]
            for i, work_dict in enumerate(work_list):
                headers += [f"Start {i+1}", f"End {i+1}"]
                row_content[0] += [work_dict["start"].strftime("%H:%M"), work_dict["end"].strftime("%H:%M") if work_dict["end"] else ""]
        
        print(tabulate(row_content,headers=headers, tablefmt="grid"))
        
    def save(self):
        self.workbook.save(EXCEL_FILE)
        